# -*- coding: utf-8 -*-
"""
dw_bms / models / invoice_import_wizard.py

Multi-step wizard for XLSX invoice import:
  Step 1  (state='upload')  — Upload file, click "Read File"
  Step 2  (state='mapping') — Review/edit XLSX-column → Odoo-field mappings, click "Import"

Import engine:
  • Reads rows using confirmed mapping (not positional)
  • Groups rows by invoice_number field
  • Skips duplicates; auto-creates partners (GSTIN-validated) and products
  • Resolves CGST+SGST vs IGST, creates sale order, confirms, invoices, posts
  • Writes dw.invoice.import.log and opens it after import
"""

import base64
import io
import logging
import re
from collections import defaultdict
from datetime import datetime, timedelta

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .invoice_import_column_map import ODOO_FIELD_SELECTION

_logger = logging.getLogger(__name__)


# ─── SYNONYM TABLE: normalised-header → canonical odoo_field ─────────────────
# Keys are already normalised (lowercase, spaces→_, stripped).
SYNONYMS = {
    # Invoice
    "invoice_no":             "invoice_number",
    "invoice_no.":            "invoice_number",
    "inv_no":                 "invoice_number",
    "bill_no":                "invoice_number",
    "date":                   "invoice_date",
    "inv_date":               "invoice_date",
    "invoice_type":           "invoice_type",
    # Customer
    "customer":               "customer_name",
    "party_name":             "customer_name",
    "client_name":            "customer_name",
    "buyer":                  "customer_name",
    "gst_number":             "customer_gstin",
    "gstin":                  "customer_gstin",
    "gst_no":                 "customer_gstin",
    "vat_number":             "customer_gstin",
    "tax_id":                 "customer_gstin",
    # Contact
    "contact":                "contact_id",
    "contact_no":             "contact_number",
    "phone":                  "contact_number",
    "mobile":                 "contact_number",
    # Product
    "product":                "product_name",
    "item_name":              "product_name",
    "description":            "product_name",
    "particulars":            "product_name",
    "sku":                    "product_sku",
    "item_code":              "product_sku",
    "product_code":           "product_sku",
    "product_location":       "product_storage_location",
    "product_locations":      "product_storage_location",
    "storage_location":       "product_storage_location",
    "hsn":                    "hsn_code",
    "hsn_sac":                "hsn_code",
    # Quantities
    "qty":                    "quantity",
    "units":                  "quantity",
    "unit":                   "unit_of_measure",
    "uom":                    "unit_of_measure",
    # Price
    "rate":                   "unit_price",
    "price":                  "unit_price",
    "unit_rate":              "unit_price",
    "discount":               "discount_percent",
    "disc%":                  "discount_percent",
    # Tax rates
    "cgst_rate":              "cgst_rate",
    "cgst%":                  "cgst_rate",
    "sgst_rate":              "sgst_rate",
    "sgst%":                  "sgst_rate",
    "igst_rate":              "igst_rate",
    "igst%":                  "igst_rate",
    "tax":                    "tax_percent",
    "gst%":                   "tax_percent",
    "tax%":                   "tax_percent",
    "tax_rate":               "tax_percent",
    "tax_percentage":         "tax_percent",
    "tax_(%)":                "tax_percent",
    "tax_(in_%)":             "tax_percent",
    # Tax amounts
    "cgst_rate_price":        "cgst_amount",
    "cgst_amount":            "cgst_amount",
    "sgst_rate_price":        "sgst_amount",
    "sgst_amount":            "sgst_amount",
    "igst_rate_prices":       "igst_amount",
    "igst_amount":            "igst_amount",
    # Totals
    "amount":                 "line_total",
    "total":                  "line_total",
    "price_inc._tax":         "price_with_tax",
    "price_inc_tax":          "price_with_tax",
    "net_total":              "grand_total",
    "invoice_total":          "grand_total",
    # E-Invoice
    "e-invoice_ack._no.":     "ack_number",
    "e-invoice_ack_no":       "ack_number",
    "e-invoice_ack":          "ack_number",
    "ack_no":                 "ack_number",
    "e-invoice_date":         "ack_date",
    "e-invoice_irn_number":   "irn_number",
    "e-invoice_irn":          "irn_number",
    "irn":                    "irn_number",
    "e-invoice_amt":          "e_invoice_amount",
    "e-invoice_amount":       "e_invoice_amount",
    # E-Way Bill
    "e-way_bill_no.":         "eway_bill_number",
    "e-way_bill_no":          "eway_bill_number",
    "eway_bill_no":           "eway_bill_number",
    "e-way_bill_date":        "eway_bill_date",
    "e-way_amt":              "eway_bill_amount",
    "e-way_amount":           "eway_bill_amount",
    # Payment
    "payment_method":         "payment_mode",
    "mode_of_payment":        "payment_mode",
    "bank":                   "bank_name",
    "ref":                    "payment_reference",
    "utr":                    "payment_reference",
    # Addresses
    "billing_address":        "billing_address",
    "bill_address":           "billing_address",
    "billing_city":           "billing_city",
    "bill_city":              "billing_city",
    "billing_state":          "billing_state",
    "bill_state":             "billing_state",
    "billing_country":        "billing_country",
    "bill_country":           "billing_country",
    "billing_pincode":        "billing_pincode",
    "billing_pin_code":       "billing_pincode",
    "billing_pin_cod":        "billing_pincode",
    "bill_pincode":           "billing_pincode",
    "bill_pin_code":          "billing_pincode",
    "bill_pin_cod":           "billing_pincode",
    "shipping_address":       "shipping_address",
    "shipping_city":          "shipping_city",
    "shipping_state":         "shipping_state",
    "shipping_country":       "shipping_country",
    "shipping_pincode":       "shipping_pincode",
    "shipping_pin_code":      "shipping_pincode",
    "shipping_pin_cod":       "shipping_pincode",
    "tax_amount":             "total_tax_amount",
    "total_tax":              "total_tax_amount",
    "currency_code":          "currency",
    "curr":                   "currency",
    "ecommerce":              "ecommerce_platform",
    "e_commerce_platform":    "ecommerce_platform",
    "platform_order":         "platform_order_id",
    "order_id":               "platform_order_id",
    "place_of_supply_state":  "place_of_supply",
    "order_type":             "invoice_type",
    "sale_type":              "invoice_type",
}

# Valid canonical field names (the stored VALUES from selection, not display labels)
VALID_FIELDS = {k for k, v in ODOO_FIELD_SELECTION if k != "skip"}

# Payment mode → journal type
PAYMENT_MODE_MAP = [
    (["cash"],                              "cash"),
    (["neft", "rtgs", "imps", "bank",
      "bank_transfer", "cheque", "dd"],      "bank"),
    (["upi", "gpay", "phonepe", "paytm",
      "bhim"],                               "bank"),
    (["credit_card", "debit_card", "card"], "bank"),
]

GSTIN_RE = re.compile(
    r"^\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}Z[A-Z\d]{1}$"
)


def _norm(raw):
    """Normalise a header: strip, lower, collapse whitespace→_."""
    return re.sub(r"\s+", "_", str(raw).strip().lower())


def _safe(val):
    if val is None:
        return ""
    try:
        import math
        if isinstance(val, float):
            if math.isnan(val):
                return ""
            # Convert whole-number floats to int strings: 1727.0 → "1727"
            if val == int(val):
                return str(int(val)).strip()
    except Exception:
        pass
    return str(val).strip()


def _to_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, "date") and callable(val.date):
        return val.date()
    if hasattr(val, "isoformat"):
        return val
    if isinstance(val, (int, float)):
        # Excel date serial fallback (when cell isn't recognized as date)
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(val))).date()
        except Exception:
            return None
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M", "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return None


def _float(val):
    try:
        s = str(val).strip().replace(",", "")
        s = s.replace("%", "").replace("₹", "").replace("$", "")
        return float(s) if s else 0.0
    except (TypeError, ValueError):
        return 0.0


def _percent(val):
    """
    Parse percent values from XLSX.
    Supports:
      - 3, 18
      - "3%", "18 %"
      - Excel percentage cells read as decimals: 0.03, 0.18
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        f = float(val)
        if 0 < abs(f) <= 1:
            return round(f * 100.0, 4)
        return round(f, 4)
    s = str(val).strip()
    has_pct_symbol = "%" in s
    f = _float(s)
    if not has_pct_symbol and 0 < abs(f) <= 1:
        return round(f * 100.0, 4)
    return round(f, 4)


def _gstin_ok(val):
    return bool(GSTIN_RE.match(_safe(val).upper().strip()))


# ─── WIZARD ──────────────────────────────────────────────────────────────────

class DwInvoiceImportWizard(models.TransientModel):
    _name = "dw.invoice.import.wizard"
    _description = "XLSX Invoice Import Wizard"

    _INVOICE_TYPE_LABEL_MAP = {
        "flipkart wb": "flipkart_wb",
        "vastu craft (delhi)": "vastu_delhi",
        "vastu craft delhi": "vastu_delhi",
        "daily sales": "daily_sales",
        "flipkart mh": "flipkart_mh",
        "kv enterprises (haryana)": "kv_hr",
        "kv enterprise (haryana)": "kv_hr",
        "website sales": "website_sales",
        "export sales": "export_sales",
        "kv enterprises (tamil nadu)": "kv_tn",
        "kv enterprises (karnataka)": "kv_ka",
        "kv enterprises (maharashtra)": "kv_mh",
        "kv enterprises (west bengal)": "kv_wb",
        "kv enterprises (telangana)": "kv_tel",
    }

    state = fields.Selection(
        selection=[("upload", "Upload"), ("mapping", "Map Fields")],
        default="upload",
        required=True,
    )
    xlsx_file = fields.Binary(string="XLSX File", attachment=False)
    xlsx_filename = fields.Char(string="File Name")
    column_map_ids = fields.One2many(
        comodel_name="dw.invoice.import.column.map",
        inverse_name="wizard_id",
        string="Column Mapping",
    )
    mapping_note = fields.Char(
        string="Note",
        readonly=True,
        default="Review the auto-detected mappings below. Change any dropdown to correct it, then click Import.",
    )

    # ── STEP 1: Read headers ──────────────────────────────────────────────────

    def action_read_headers(self):
        """Parse XLSX row 0, create column_map_ids with auto-detected odoo_field."""
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError(_("Please upload an XLSX file first."))

        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl is required: pip install openpyxl"))

        raw = base64.b64decode(self.xlsx_file)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active

        headers = []
        sample_row = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [_safe(c) for c in row]
            elif row_idx == 1:
                sample_row = [_safe(c) for c in row]
                break

        wb.close()

        if not any(headers):
            raise UserError(_("Row 1 of the XLSX is empty — no column headers found."))

        # Remove old mappings
        self.column_map_ids.unlink()

        map_records = []
        for seq, raw_header in enumerate(headers, start=1):
            if not raw_header:
                continue
            normalised = _norm(raw_header)
            # Auto-detect odoo_field
            if normalised in VALID_FIELDS:
                odoo_field = normalised
            elif normalised in SYNONYMS and SYNONYMS[normalised] in VALID_FIELDS:
                odoo_field = SYNONYMS[normalised]
            else:
                odoo_field = "skip"

            col_idx_0 = seq - 1  # real 0-based column position in xlsx
            sample = sample_row[col_idx_0] if col_idx_0 < len(sample_row) else ""
            map_records.append({
                "wizard_id": self.id,
                "sequence": seq,
                "col_index": col_idx_0,
                "xlsx_column": raw_header,
                "sample_value": sample[:60] if sample else "",
                "odoo_field": odoo_field,
            })

        self.env["dw.invoice.import.column.map"].create(map_records)
        self.state = "mapping"

        # Return the same wizard (re-open to show step 2)
        return {
            "type": "ir.actions.act_window",
            "res_model": "dw.invoice.import.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "views": [(False, "form")],
            "target": "new",
        }

    def action_back(self):
        """Go back to upload step."""
        self.state = "upload"
        self.column_map_ids.unlink()
        return {
            "type": "ir.actions.act_window",
            "res_model": "dw.invoice.import.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "views": [(False, "form")],
            "target": "new",
        }

    # ── STEP 2: Import ────────────────────────────────────────────────────────

    def action_import(self):
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError(_("No file uploaded."))
        if not self.column_map_ids:
            raise UserError(_("No column mappings defined. Please go back and read the file first."))

        # Build {odoo_field: actual_xlsx_col_index} from confirmed mappings
        field_to_col = {}
        duplicate_mapped_fields = set()
        for mapping in self.column_map_ids:
            if mapping.odoo_field and mapping.odoo_field != "skip":
                if mapping.odoo_field in field_to_col:
                    duplicate_mapped_fields.add(mapping.odoo_field)
                field_to_col[mapping.odoo_field] = mapping.col_index  # real 0-based xlsx column

        if duplicate_mapped_fields:
            _logger.warning(
                "Duplicate mappings detected for fields %s. "
                "Continuing import with last-mapped column for each field.",
                ", ".join(sorted(duplicate_mapped_fields)),
            )

        _logger.info("Invoice import field_to_col: %s", field_to_col)

        if "invoice_number" not in field_to_col:
            raise UserError(_(
                "No column is mapped to 'Invoice Number'.\n"
                "Please go back to Step 2 and map your invoice number column."
            ))

        rows = self._parse_xlsx(field_to_col)
        if not rows:
            raise UserError(_("No data rows found in the uploaded file."))

        grouped = defaultdict(list)
        for row in rows:
            inv_no = _safe(row.get("invoice_number"))
            if inv_no:
                grouped[inv_no].append(row)

        if not grouped:
            raise UserError(
                _("Could not find any invoice numbers. "
                  "Make sure the 'Invoice Number' column is correctly mapped.")
            )

        log_lines = []
        created = skipped = failed = 0

        for inv_number, inv_rows in grouped.items():
            try:
                with self.env.cr.savepoint():
                    result = self._process_invoice(inv_number, inv_rows)
                    if result["status"] == "created":
                        created += 1
                    elif result["status"] == "skipped":
                        skipped += 1
                    log_lines.append(result)
            except Exception as exc:
                failed += 1
                _logger.exception("Import failed [%s]: %s", inv_number, exc)
                log_lines.append({
                    "invoice_number": inv_number,
                    "status": "failed",
                    "message": str(exc),
                    "move_id": False,
                })

        total = created + skipped + failed
        log = self.env["dw.invoice.import.log"].create({
            "filename": self.xlsx_filename or "unknown.xlsx",
            "total_invoices": total,
            "created": created,
            "skipped": skipped,
            "failed": failed,
            "state": "done" if failed == 0 else "partial",
            "log_line_ids": [
                (0, 0, {
                    "invoice_number": ll["invoice_number"],
                    "status": ll["status"],
                    "message": ll.get("message", ""),
                    "move_id": ll.get("move_id", False),
                })
                for ll in log_lines
            ],
        })

        for line in log.log_line_ids:
            if line.move_id and line.status == "created":
                line.move_id.sudo().write({"import_log_line_id": line.id})

        return {
            "type": "ir.actions.act_window",
            "name": _("Import Log — %s", log.name),
            "res_model": "dw.invoice.import.log",
            "res_id": log.id,
            "view_mode": "form",
            "views": [(False, "form")],
            "target": "current",
        }

    # ── XLSX PARSING ──────────────────────────────────────────────────────────

    def _parse_xlsx(self, field_to_col):
        """Return list of dicts keyed by canonical odoo_field names."""
        import openpyxl
        raw = base64.b64decode(self.xlsx_file)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue  # skip header
            if not any(row):
                continue  # skip blank rows
            row_data = {}
            for field_name, col_idx in field_to_col.items():
                row_data[field_name] = row[col_idx] if col_idx < len(row) else None
            inv_no = _safe(row_data.get("invoice_number"))
            if not inv_no:
                continue
            row_data["invoice_number"] = inv_no
            rows.append(row_data)
        wb.close()
        return rows

    # ── INVOICE PROCESSING ────────────────────────────────────────────────────

    def _process_invoice(self, inv_number, rows):
        header = rows[0]

        invoice_type = self._map_invoice_type(header.get("invoice_type"))

        # Duplicate check
        existing = self.env["account.move"].sudo().search(
            [("ref", "=", inv_number), ("move_type", "=", "out_invoice"), ("company_id", "=", self.env.company.id)], limit=1
        )
        if existing:
            return {
                "invoice_number": inv_number,
                "status": "skipped",
                "message": f"Already exists: {existing.name} (ID {existing.id})",
                "move_id": existing.id,
            }

        partner = self._get_or_create_partner(header)
        currency = self._get_currency(header.get("currency"))

        company_state = self.env.company.partner_id.state_id or self.env.company.state_id
        bill_state_id = self._get_state_id(header.get("billing_state"), header.get("billing_country"))
        ship_state_id = self._get_state_id(header.get("shipping_state"), header.get("shipping_country"))
        effective_state_id = bill_state_id or ship_state_id or partner.state_id.id
        is_intra = bool(company_state and effective_state_id and company_state.id == effective_state_id)

        order_line_vals = []
        for row in rows:
            line = self._build_sale_order_line(row, is_intra, partner)
            if line:
                order_line_vals.append((0, 0, line))

        if not order_line_vals:
            raise UserError(f"No valid product lines for invoice '{inv_number}'.")

        fiscal = self._get_fiscal_position(is_intra)
        inv_date = _to_date(header.get("invoice_date"))
        sale_order = self.env["sale.order"].sudo().with_context(skip_stock_validation=True).create({
            "partner_id": partner.id,
            "date_order": inv_date or fields.Datetime.now(),
            "currency_id": currency.id if currency else self.env.company.currency_id.id,
            "fiscal_position_id": fiscal.id if fiscal else False,
            "client_order_ref": inv_number,
            "order_line": order_line_vals,
        })
        sale_order.with_context(skip_stock_validation=True).action_confirm()
        invoice = sale_order.with_context(skip_stock_validation=True)._create_invoices()
        invoice = invoice.filtered(lambda move: move.move_type == "out_invoice")[:1]
        if not invoice:
            raise UserError(f"Unable to create a customer invoice for sales order '{sale_order.name}'.")

        move_vals = {
            "ref": inv_number,
            "invoice_date": inv_date,
            "fiscal_position_id": fiscal.id if fiscal else False,
            "invoice_type": invoice_type,
            # Address fields
            "bill_to_same_as_customer": False,
            "ship_to_same_as_customer": False,
            "bill_to_address": _safe(header.get("billing_address")) or False,
            "bill_to_city":    _safe(header.get("billing_city")) or False,
            "bill_to_state_id": self._get_state_id(header.get("billing_state"), header.get("billing_country")),
            "bill_to_zip":     _safe(header.get("billing_pincode")) or False,
            "bill_to_country": _safe(header.get("billing_country")) or False,
            "ship_to_address": _safe(header.get("shipping_address")) or False,
            "ship_to_city":    _safe(header.get("shipping_city")) or False,
            "ship_to_state_id": self._get_state_id(header.get("shipping_state"), header.get("shipping_country")),
            "ship_to_zip":     _safe(header.get("shipping_pincode")) or False,
            "ship_to_country": _safe(header.get("shipping_country")) or False,
            # Customer / contact
            "dw_customer_gstin":   _safe(header.get("customer_gstin")) or False,
            "dw_contact_id":       _safe(header.get("contact_id")) or False,
            "dw_contact_number":   _safe(header.get("contact_number")) or False,
            # E-Invoice
            "dw_irn_number":       _safe(header.get("irn_number")) or False,
            "dw_ack_number":       _safe(header.get("ack_number")) or False,
            "dw_ack_date":         _to_date(header.get("ack_date")),
            "dw_e_invoice_amount": _float(header.get("e_invoice_amount")),
            # E-Way Bill
            "dw_eway_bill_number": _safe(header.get("eway_bill_number")) or False,
            "dw_eway_bill_date":   _to_date(header.get("eway_bill_date")),
            "dw_eway_bill_amount": _float(header.get("eway_bill_amount")),
            # Payment
            "dw_payment_mode":        _safe(header.get("payment_mode")) or False,
            "dw_bank_name":           _safe(header.get("bank_name")) or False,
            "dw_payment_reference":   _safe(header.get("payment_reference")) or False,
            "dw_payment_date_imported": _to_date(header.get("payment_date")),
            # Misc
            "dw_place_of_supply":      _safe(header.get("place_of_supply")) or False,
            "dw_ecommerce_platform":   _safe(header.get("ecommerce_platform")) or False,
            "dw_platform_order_id":    _safe(header.get("platform_order_id")) or False,
            "dw_grand_total_imported": _float(header.get("grand_total")),
        }
        invoice.sudo().write(move_vals)
        invoice.action_post()

        return {
            "invoice_number": inv_number,
            "status": "created",
            "message": f"Sale Order {sale_order.name} confirmed and invoice {invoice.name} posted.",
            "move_id": invoice.id,
        }

    # ── LINE BUILDER ──────────────────────────────────────────────────────────

    def _build_sale_order_line(self, row, is_intra, partner):
        product, product_uom_id = self._get_or_create_product(row)
        qty = _float(row.get("quantity")) or 1.0
        price = _float(row.get("unit_price"))
        disc = _percent(row.get("discount_percent"))

        # Determine tax% — prefer explicit cgst+sgst/igst, fall back to tax_percent
        tax_pct = _percent(row.get("tax_percent"))
        cgst_r = _percent(row.get("cgst_rate"))
        sgst_r = _percent(row.get("sgst_rate"))
        igst_r = _percent(row.get("igst_rate"))
        if not tax_pct:
            if is_intra and (cgst_r or sgst_r):
                tax_pct = cgst_r + sgst_r
            elif igst_r:
                tax_pct = igst_r
            else:
                taxable_value = _float(row.get("taxable_value"))
                total_tax_amount = _float(row.get("total_tax_amount"))
                if taxable_value > 0 and total_tax_amount > 0:
                    tax_pct = (total_tax_amount / taxable_value) * 100.0

        tax_ids = self._get_taxes(tax_pct, is_intra)
        price_with_tax = _float(row.get("price_with_tax"))
        if not price_with_tax:
            tax_compute = tax_ids.compute_all(
                price,
                currency=self.env.company.currency_id,
                quantity=1.0,
                product=product,
                partner=partner,
            )
            price_with_tax = tax_compute.get("total_included", price)

        line_vals = {
            "product_id": product.id,
            "name": _safe(row.get("product_name")) or product.name,
            "product_uom_qty": qty,
            "price_unit": price,
            "discount": disc,
            "tax_id": [(6, 0, tax_ids.ids)],
            "price_incl_tax": price_with_tax,
        }

        if product_uom_id:
            line_vals["product_uom"] = product_uom_id

        return line_vals

    # ── HELPERS ───────────────────────────────────────────────────────────────

    def _get_or_create_partner(self, header):
        Partner = self.env["res.partner"].sudo()
        gstin = _safe(header.get("customer_gstin"))
        name = _safe(header.get("customer_name"))
        company_id = self.env.company.id
        if not name:
            raise UserError(_("Customer name is missing."))
        if gstin and _gstin_ok(gstin):
            p = Partner.search(
                [("vat", "=", gstin.upper()), ("company_id", "in", [company_id, False])],
                limit=1
            )
            if p:
                return p
        p = Partner.search(
            [("name", "=ilike", name), ("customer_rank", ">", 0), ("company_id", "in", [company_id, False])],
            limit=1
        )
        if p:
            return p
        vals = {"name": name, "customer_rank": 1}
        if gstin and _gstin_ok(gstin):
            vals["vat"] = gstin.upper()
        country = self._get_country(header.get("billing_country"))
        state_id = self._get_state_id(header.get("billing_state"), header.get("billing_country"))
        if country:
            vals["country_id"] = country.id
        if state_id:
            vals["state_id"] = state_id
        vals["zip"] = _safe(header.get("billing_pincode")) or False
        vals["street"] = _safe(header.get("billing_address")) or False
        if _safe(header.get("contact_number")):
            vals["phone"] = _safe(header.get("contact_number"))
        vals["company_id"] = company_id
        return Partner.create(vals)

    def _get_or_create_product(self, row):
        PP = self.env["product.product"].sudo()
        name = _safe(row.get("product_name")) or _safe(row.get("product_sku")) or "Imported Product"
        product_location = _safe(row.get("product_storage_location"))
        hsn = _safe(row.get("hsn_code"))
        uom_name = _safe(row.get("unit_of_measure"))

        # Lookup UoM first
        product_uom_id = False
        if uom_name:
            norm_uom_name = re.sub(r"[^a-z0-9]+", "", uom_name.lower().strip())
            uom_map = {
                "pcs": "Units", "nos": "Units", "pc": "Units", "pieces": "Units",
                "unit": "Units", "service": "Units", "srv": "Units",
                "kg": "kg", "kgs": "kg", "g": "g", "gm": "g", "gms": "g",
                "mtr": "m", "meter": "m", "l": "L", "ltr": "L", "liters": "L",
            }
            search_uom = uom_map.get(norm_uom_name, uom_name)
            uom = self.env["uom.uom"].sudo().search([("name", "=ilike", search_uom)], limit=1)
            if not uom and search_uom != uom_name:
                uom = self.env["uom.uom"].sudo().search([("name", "=ilike", uom_name)], limit=1)
            if uom:
                product_uom_id = uom.id

        def _update_product_uom(product_variant):
            if product_uom_id and product_variant.uom_id.id != product_uom_id:
                # Update but only if categories don't clash, or if we force it via sudo (which we are)
                # But to avoid UserError from Odoo when changing uom on products with stock, we try/except
                try:
                    product_variant.product_tmpl_id.sudo().write({
                        "uom_id": product_uom_id,
                        "uom_po_id": product_uom_id
                    })
                except Exception as e:
                    # If Odoo blocks the uom change (e.g. established stock history),
                    # we must return the original product's uom so the invoice line matches it
                    return product_variant.uom_id.id
            return product_uom_id

        p = PP.search([("name", "=ilike", name)], limit=1)
        if p:
            if product_location:
                p.product_tmpl_id.sudo().write({"product_storage_location": product_location})
            final_uom_id = _update_product_uom(p)
            return p, final_uom_id

        # Auto-create
        tmpl_vals = {"name": name, "default_code": False, "type": "consu", "sale_ok": True}
        if hsn:
            tmpl_vals["l10n_in_hsn_code"] = hsn
        if product_location:
            tmpl_vals["product_storage_location"] = product_location
        if product_uom_id:
            tmpl_vals["uom_id"] = product_uom_id
            tmpl_vals["uom_po_id"] = product_uom_id

        tmpl = self.env["product.template"].sudo().create(tmpl_vals)
        return tmpl.product_variant_ids[0], product_uom_id

    def _get_taxes(self, tax_pct, is_intra):
        Tax = self.env["account.tax"].sudo()
        TaxGroup = self.env["account.tax.group"].sudo()
        company_id = self.env.company.id
        if not tax_pct:
            return Tax.browse([])
        base = [("type_tax_use", "=", "sale"), ("company_id", "=", company_id), ("active", "=", True)]

        def _pick_by_amount(candidates, amount):
            for tax in candidates:
                if abs((tax.amount or 0.0) - amount) < 1e-6:
                    return tax
            return Tax.browse([])

        def _tax_group_id(group_name):
            grp = TaxGroup.search(
                [("company_id", "=", company_id), ("name", "=ilike", group_name)],
                limit=1
            )
            return grp.id if grp else False

        def _get_or_create_tax(name, amount, group_name):
            candidates = Tax.search(base + [("name", "ilike", name), ("tax_group_id.name", "=ilike", group_name)])
            tax = _pick_by_amount(candidates, amount)
            if tax:
                return tax
            return Tax.create({
                "name": name,
                "amount_type": "percent",
                "amount": amount,
                "type_tax_use": "sale",
                "company_id": company_id,
                "tax_group_id": _tax_group_id(group_name),
                "active": True,
            })

        if is_intra:
            half = tax_pct / 2
            cgst = _get_or_create_tax(f"CGST {half:g}%", half, "CGST")
            sgst = _get_or_create_tax(f"SGST {half:g}%", half, "SGST")
            return cgst | sgst
        else:
            igst = _get_or_create_tax(f"IGST {tax_pct:g}%", tax_pct, "IGST")
            return igst

    def _get_fiscal_position(self, is_intra):
        name = "GST Intra State" if is_intra else "GST Inter State"
        return self.env["account.fiscal.position"].sudo().search(
            [("name", "=", name), ("company_id", "=", self.env.company.id)], limit=1
        )

    def _map_invoice_type(self, raw_value):
        value = _safe(raw_value)
        if not value:
            raise UserError(_("Invoice Type is required in the import file."))

        normalized = value.strip().lower()
        invoice_type_values = self.env["account.move"]._fields["invoice_type"].selection
        valid_values = {key for key, _label in invoice_type_values}
        if normalized in valid_values:
            return normalized

        mapped_value = self._INVOICE_TYPE_LABEL_MAP.get(normalized)
        if mapped_value:
            return mapped_value

        raise UserError(_("Invalid Invoice Type '%s' in the import file.") % value)

    def _get_currency(self, name):
        n = _safe(name)
        return self.env["res.currency"].sudo().search(
            [("name", "=ilike", n), ("active", "in", [True, False])], limit=1
        ) if n else None

    def _get_country(self, name):
        n = _safe(name)
        return self.env["res.country"].sudo().search([("name", "=ilike", n)], limit=1) if n else None

    def _get_state_id(self, state_name, country_name=None):
        n = _safe(state_name)
        if not n:
            return False
        domain = [("name", "=ilike", n)]
        if country_name:
            c = self._get_country(country_name)
            if c:
                domain.append(("country_id", "=", c.id))
        state = self.env["res.country.state"].sudo().search(domain, limit=1)
        return state.id if state else False
