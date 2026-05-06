import base64
import io

import openpyxl

from odoo import api, fields, models
from odoo.exceptions import UserError
from odoo.tools.float_utils import float_compare


class InvoiceImportSheet(models.TransientModel):
    _name = 'dw.invoice.import.v2.sheet'
    _description = 'DW Import Sheet'

    name = fields.Char(string='Sheet Name', required=True)


class InvoiceImportWizard(models.TransientModel):
    _name = 'dw.invoice.import.v2.wizard'
    _description = 'DW SKU Invoice Import Wizard'

    file = fields.Binary(required=True)
    available_sheet_ids = fields.Many2many('dw.invoice.import.v2.sheet', string="Available Sheets")
    sheet_id = fields.Many2one('dw.invoice.import.v2.sheet', string='Sheet', required=True)
    partner_id = fields.Many2one('res.partner', string='Customer', required=True)
    invoice_date = fields.Date(required=True)
    shipping_id = fields.Char(string='Shipping Id')
    invoice_type = fields.Selection(
        selection=lambda self: self.env['account.move']._fields['invoice_type'].selection,
        required=True,
    )

    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done'),
    ], default='draft')

    @staticmethod
    def _normalize_key(value):
        return str(value).strip().lower() if value not in (False, None) else False

    def _decode_workbook(self):
        if not self.file:
            raise UserError("Please upload file.")
        try:
            data = base64.b64decode(self.file)
            return openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        except Exception as error:
            raise UserError("Please upload a valid Excel file.") from error

    @api.onchange('file')
    def _onchange_file(self):
        self.sheet_id = False
        self.available_sheet_ids = [(5, 0, 0)]
        if not self.file:
            return

        wb = self._decode_workbook()
        sheet_names = wb.sheetnames
        if not sheet_names:
            raise UserError("No sheets found in the uploaded Excel file.")

        sheets = self.env['dw.invoice.import.v2.sheet'].create([
            {'name': name} for name in sheet_names
        ])
        
        self.available_sheet_ids = [(6, 0, sheets.ids)]
        self.sheet_id = sheets[0] if sheets else False

    def _get_outgoing_picking_type(self, company):
        picking_type = self.env['stock.picking.type'].search([
            ('code', '=', 'outgoing'),
            ('warehouse_id.company_id', '=', company.id),
        ], order='sequence, id', limit=1)
        if not picking_type:
            picking_type = self.env['stock.picking.type'].search([
                ('code', '=', 'outgoing'),
                ('company_id', 'in', [company.id, False]),
            ], order='company_id desc, sequence, id', limit=1)
        if not picking_type:
            raise UserError("No outgoing operation type found for the invoice company.")
        return picking_type

    def _get_invoice_fiscal_position(self, partner, company):
        company_state = company.partner_id.state_id or company.state_id
        partner_state = partner.state_id

        if not company_state or not partner_state:
            return self.env['account.fiscal.position']

        fiscal_position_names = [
            'GST Intra State',
            'Intra State',
        ] if company_state.id == partner_state.id else [
            'GST Inter State',
            'Inter State',
        ]
        fiscal_position = self.env['account.fiscal.position'].search([
            ('name', 'in', fiscal_position_names),
            ('company_id', '=', company.id),
        ], limit=1)
        if fiscal_position:
            return fiscal_position

        # Fallback to any company fiscal position that auto-detects this state rule.
        state_domain = [('state_ids', 'in', partner_state.id)]
        if company_state.id == partner_state.id:
            state_domain = [('state_ids', 'in', partner_state.id)]
        else:
            state_domain = [('state_ids', 'in', partner_state.id)]

        return self.env['account.fiscal.position'].search([
            ('company_id', '=', company.id),
            ('auto_apply', '=', True),
            ('country_id.code', '=', 'IN'),
            *state_domain,
        ], limit=1)

    def _check_stock_availability(self, product_qty, company):
        stock_products = self.env['product.product'].browse(
            [product_id for product_id, qty in product_qty.items() if qty]
        ).exists().filtered(lambda product: product.detailed_type == 'product')

        insufficient_products = []
        for product in stock_products.with_company(company):
            required_qty = product_qty.get(product.id, 0.0)
            rounding = product.uom_id.rounding or 0.01
            if float_compare(product.qty_available, required_qty, precision_rounding=rounding) < 0:
                insufficient_products.append(
                    "%s: required %.2f, available %.2f"
                    % (product.display_name, required_qty, product.qty_available)
                )

        if insufficient_products:
            raise UserError(
                "Stock is not available for the following product(s):\n%s"
                % "\n".join(insufficient_products[:10])
            )

        return stock_products

    def _create_and_validate_delivery(self, invoice, product_qty, stock_products):
        if not stock_products:
            return self.env['stock.picking']

        picking_type = self._get_outgoing_picking_type(invoice.company_id)
        source_location = picking_type.default_location_src_id
        dest_location = (
            picking_type.default_location_dest_id
            or self.env.ref('stock.stock_location_customers', raise_if_not_found=False)
        )

        if not source_location or not dest_location:
            raise UserError("Source or destination location is missing for outgoing delivery.")

        delivery_partner = invoice.shipping_partner_id or invoice.partner_id
        picking = self.env['stock.picking'].create({
            'partner_id': delivery_partner.id,
            'origin': invoice.name,
            'picking_type_id': picking_type.id,
            'location_id': source_location.id,
            'location_dest_id': dest_location.id,
            'company_id': invoice.company_id.id,
        })

        move_vals_list = []
        for product in stock_products:
            qty = product_qty.get(product.id)
            if not qty:
                continue
            move_vals_list.append({
                'name': product.display_name,
                'picking_id': picking.id,
                'product_id': product.id,
                'product_uom_qty': qty,
                'product_uom': product.uom_id.id,
                'location_id': source_location.id,
                'location_dest_id': dest_location.id,
                'company_id': invoice.company_id.id,
            })

        if not move_vals_list:
            return self.env['stock.picking']

        self.env['stock.move'].create(move_vals_list)
        picking.action_confirm()
        picking.action_assign()

        unreserved_moves = picking.move_ids_without_package.filtered(
            lambda move: move.state not in ('assigned', 'done', 'cancel')
        )
        if unreserved_moves:
            raise UserError(
                "Delivery could not be reserved for: %s"
                % ", ".join(unreserved_moves.mapped('product_id.display_name'))
            )

        moves_to_validate = picking.move_ids_without_package.filtered(
            lambda move: move.state not in ('done', 'cancel')
        )
        moves_to_validate.write({
            'quantity': 0.0,
            'picked': True,
        })
        for move in moves_to_validate:
            move.quantity = move.product_uom_qty

        picking.button_validate()
        return picking

    def action_import(self):
        wb = self._decode_workbook()
        if self.sheet_id and self.sheet_id.name in wb.sheetnames:
            sheet = wb[self.sheet_id.name]
        else:
            sheet = wb.active

        alias_model = self.env['dw.product.name.alias']
        product_model = self.env['product.product']

        alias_map = {
            self._normalize_key(alias.name): alias.product_tmpl_id.product_variant_id.id
            for alias in alias_model.search([])
            if alias.name and alias.product_tmpl_id.product_variant_id
        }
        product_name_map = {
            self._normalize_key(product['name']): product['id']
            for product in product_model.search_read([], ['name'])
            if product.get('name')
        }

        product_qty = {}
        errors = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            sku = row[0] if len(row) > 0 else False
            qty = row[1] if len(row) > 1 else False
            name = row[2] if len(row) > 2 else False

            if not sku or not qty:
                continue

            product_id = (
                alias_map.get(self._normalize_key(sku))
                or product_name_map.get(self._normalize_key(sku))
                or product_name_map.get(self._normalize_key(name))
            )

            if not product_id:
                identifier = name.strip() if name else str(sku).strip()
                errors.append(f"Product '{identifier}' is not available in the system.")
                continue

            product_qty[product_id] = product_qty.get(product_id, 0.0) + qty

        if errors:
            raise UserError("\n".join(errors[:5]))

        stock_products = self._check_stock_availability(product_qty, self.env.company)

        seq_code = f"dw.invoice.{self.invoice_type}"
        seq = self.env['ir.sequence'].search([('code', '=', seq_code)], limit=1)
        fiscal_position = self._get_invoice_fiscal_position(self.partner_id, self.env.company)

        invoice_vals = {
            'move_type': 'out_invoice',
            'partner_id': self.partner_id.id,
            'invoice_date': self.invoice_date,
            'invoice_type': self.invoice_type,
            'dw_shipping_id': self.shipping_id,
            'fiscal_position_id': fiscal_position.id,
        }

        if seq:
            invoice_vals['name'] = seq.next_by_id()

        invoice = self.env['account.move'].create(invoice_vals)
        products = product_model.browse(list(product_qty)).exists()
        products_by_id = {product.id: product for product in products}

        lines = []
        for product_id, qty in product_qty.items():
            product = products_by_id[product_id]
            product_taxes = product.taxes_id.filtered(
                lambda tax: tax.type_tax_use == 'sale' and tax.company_id == invoice.company_id
            )
            mapped_taxes = (
                fiscal_position.map_tax(product_taxes)
                if fiscal_position
                else product_taxes
            )
            lines.append((0, 0, {
                'product_id': product.id,
                'quantity': qty,
                'price_unit': product.lst_price,
                'tax_ids': [(6, 0, mapped_taxes.ids)],
            }))

        invoice.write({'invoice_line_ids': lines})
        invoice.action_post()
        picking = self._create_and_validate_delivery(invoice, product_qty, stock_products)

        self.state = 'done'

        message = f"Invoice {invoice.name} created and posted successfully."
        if picking:
            message += f" Delivery {picking.name} was created and validated."

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Invoice Created',
                'message': message,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }
